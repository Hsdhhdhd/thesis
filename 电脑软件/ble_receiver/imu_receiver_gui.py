import sys
import asyncio
import struct
import time
import queue
import math
from datetime import datetime
from collections import deque
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QComboBox, QGroupBox, QGridLayout,
    QCheckBox, QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QPointF, QTimer
from PyQt5.QtGui import QColor, QFont, QPen, QPainter
from PyQt5.QtChart import QChart, QChartView, QLineSeries, QScatterSeries
import bleak
from bleak import BleakClient, BleakScanner
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D


# UUIDs match Bluetooth device
IMU_SERVICE_UUID = "12345678-1234-5678-1234-56789abcdef0"
IMU_DATA_CHAR_UUID = "12345678-1234-5678-1234-56789abcdef1"
BLE_OUTPUT_HZ = 60.0
BLE_OUTPUT_INTERVAL_S = 1.0 / BLE_OUTPUT_HZ
IMU_SENSOR_ODR_HZ = 104.0

# Quaternion Filter (Madgwick Simplified)
class QuaternionFilter:
    """Perform quaternion integration using acceleration and angular velocity"""
    def __init__(self, beta=0.1):
        """
        beta: Filter gain (0-1), higher trusts acceleration more
        """
        self.q = [1.0, 0.0, 0.0, 0.0]  # [w, x, y, z]
        self.beta = beta
        self.default_dt = BLE_OUTPUT_INTERVAL_S
        self.g = 9.81   # Gravitational acceleration (m/s²)
        self.initialized = False
        
        # Initialization buffer: collect first N frames for averaging
        self.init_buffer = []
        self.init_frames_needed = 10  # Collect 10 frames (80ms) for averaging
        
    def quat_mult(self, q1, q2):
        """Quaternion multiplication: q1 * q2"""
        w1, x1, y1, z1 = q1
        w2, x2, y2, z2 = q2
        
        w = w1*w2 - x1*x2 - y1*y2 - z1*z2
        x = w1*x2 + x1*w2 + y1*z2 - z1*y2
        y = w1*y2 - x1*z2 + y1*w2 + z1*x2
        z = w1*z2 + x1*y2 - y1*x2 + z1*w2
        
        return [w, x, y, z]
    
    def quat_conj(self, q):
        """Quaternion conjugate: q* = [w, -x, -y, -z]"""
        return [q[0], -q[1], -q[2], -q[3]]
    
    def get_rotation_matrix(self):
        """
        Get rotation matrix corresponding to quaternion (3x3)
        Returns numpy array
        """
        q0, q1, q2, q3 = self.q
        
        # Rotation matrix calculated from quaternion
        R = np.array([
            [1 - 2*(q2**2 + q3**2), 2*(q1*q2 - q0*q3), 2*(q1*q3 + q0*q2)],
            [2*(q1*q2 + q0*q3), 1 - 2*(q1**2 + q3**2), 2*(q2*q3 - q0*q1)],
            [2*(q1*q3 - q0*q2), 2*(q2*q3 + q0*q1), 1 - 2*(q1**2 + q2**2)]
        ])
        return R
    
    def rotate_vector_by_quat(self, v, q):
        """
        Rotate vector v by quaternion
        v: [vx, vy, vz]
        q: [w, x, y, z]
        Returns rotated vector [vx', vy', vz']
        """
        v_quat = [0, v[0], v[1], v[2]]
        q_conj = self.quat_conj(q)
        
        # v' = q * v * q^(-1)
        temp = self.quat_mult(q, v_quat)
        result = self.quat_mult(temp, q_conj)
        
        return [result[1], result[2], result[3]]
    
    def init_from_accel(self, ax, ay, az):
        """
        Initialize quaternion based on acceleration direction
        Quaternion q represents rotation from Earth to sensor frame (world → sensor)
        q rotates [0, 0, 1] to [ax, ay, az] direction
        """
        # Normalize acceleration
        accel_norm = math.sqrt(ax**2 + ay**2 + az**2)
        if accel_norm < 0.1:
            return
        
        ax /= accel_norm
        ay /= accel_norm
        az /= accel_norm
        
        # Gravity direction in Earth coordinate system (downward is positive)
        # Note: accelerometer measures upward force, so [0,0,1] corresponds to [0,0,g]
        gx, gy, gz = 0, 0, 1
        
        # Calculate rotation from [0,0,1] to [ax,ay,az]
        # Rotation axis (cross product)：[0,0,1] × [ax,ay,az]
        rx = gy * az - gz * ay  # 0*az - 1*ay = -ay
        ry = gz * ax - gx * az  # 1*ax - 0*az = ax
        rz = gx * ay - gy * ax  # 0*ay - 0*ax = 0
        
        # Magnitude of rotation axis
        r_norm = math.sqrt(rx**2 + ry**2 + rz**2)
        
        if r_norm < 0.001:
            # Acceleration already points to Z-axis, no rotation needed
            # Need to determine same or opposite direction
            if az > 0:
                self.q = [1.0, 0.0, 0.0, 0.0]
            else:
                # 180 degree rotation around X-axis
                self.q = [0.0, 1.0, 0.0, 0.0]
        else:
            # Dot product (for calculating rotation angle)
            dot = gx * ax + gy * ay + gz * az
            angle = math.acos(max(-1, min(1, dot)))
            
            # Quaternion representation: rotate around axis [rx,ry,rz] by angle
            half_angle = angle / 2
            sin_half = math.sin(half_angle)
            
            self.q = [
                math.cos(half_angle),
                rx / r_norm * sin_half,
                ry / r_norm * sin_half,
                rz / r_norm * sin_half
            ]
        
        self.initialized = True
    
    def remove_gravity(self, accel):
        """
        Remove gravity component from accelerometer to get motion acceleration
        accel: [ax, ay, az] - Accelerometer reading (m/s²)
        Returns: [ax_motion, ay_motion, az_motion] - Gravity-removed acceleration
        
        Quaternion q represents rotation from world to sensor (world → sensor)
        Use 3rd column of rotation matrix to calculate gravity in sensor frame
        """
        q0, q1, q2, q3 = self.q
        
        # 3rd column of rotation matrix R_ws = q rotated [0,0,1]
        # Gravity direction in sensor frame (normalized)
        gx = 2.0 * (q1 * q3 - q0 * q2)
        gy = 2.0 * (q2 * q3 + q0 * q1)
        gz = q0**2 - q1**2 - q2**2 + q3**2
        
        # Gravity vector in sensor coordinate system (multiplied by gravitational acceleration)
        gravity_sensor = [gx * self.g, gy * self.g, gz * self.g]
        
        # Subtract gravity component from accelerometer reading
        accel_motion = [
            accel[0] - gravity_sensor[0],
            accel[1] - gravity_sensor[1],
            accel[2] - gravity_sensor[2]
        ]
        
        return accel_motion
        
    def update(self, ax, ay, az, gx, gy, gz, dt=None):
        """
        Update quaternion using Madgwick filter
        ax, ay, az: Acceleration (m/s²)
        gx, gy, gz: Angular velocity (°/s)
        Returns: ([w, x, y, z], is_initialized)
        """
        dt = dt if dt is not None else self.default_dt

        # Initialization phase: collect acceleration data and average
        if not self.initialized:
            self.init_buffer.append([ax, ay, az])
            
            if len(self.init_buffer) >= self.init_frames_needed:
                # Average acceleration
                avg_ax = sum(d[0] for d in self.init_buffer) / len(self.init_buffer)
                avg_ay = sum(d[1] for d in self.init_buffer) / len(self.init_buffer)
                avg_az = sum(d[2] for d in self.init_buffer) / len(self.init_buffer)
                
                # Initialize with average acceleration
                self.init_from_accel(avg_ax, avg_ay, avg_az)
                self.init_buffer.clear()
            
            # No fusion during initialization phase, just return current quaternion
            return self.q, self.initialized
        
        # Convert units: angular velocity from °/s to rad/s
        gx_rad = math.radians(gx)
        gy_rad = math.radians(gy)
        gz_rad = math.radians(gz)
        
        q0, q1, q2, q3 = self.q
        
        # Normalize acceleration
        accel_norm = math.sqrt(ax**2 + ay**2 + az**2)
        if accel_norm < 0.01:  # Prevent division by zero
            return self.q, self.initialized
        
        ax_norm = ax / accel_norm
        ay_norm = ay / accel_norm
        az_norm = az / accel_norm
        
        # Expected gravity direction (based on current quaternion, 3rd column of rotation matrix)
        # q 
        gx_pred = 2.0 * (q1 * q3 - q0 * q2)
        gy_pred = 2.0 * (q0 * q1 + q2 * q3)
        gz_pred = q0*q0 - q1*q1 - q2*q2 + q3*q3
        
        # Objective function: f = predicted - actual
        f1 = gx_pred - ax_norm
        f2 = gy_pred - ay_norm
        f3 = gz_pred - az_norm
        
        # Jacobian matrix J (3x4), each row is partial derivative of f_i wrt [q0,q1,q2,q3]
        # df1/dq = [-2*q2, 2*q3, -2*q0, 2*q1]
        # df2/dq = [2*q1, 2*q0, 2*q3, 2*q2]
        # df3/dq = [2*q0, -2*q1, -2*q2, 2*q3]
        j11, j12, j13, j14 = -2*q2, 2*q3, -2*q0, 2*q1
        j21, j22, j23, j24 = 2*q1, 2*q0, 2*q3, 2*q2
        j31, j32, j33, j34 = 2*q0, -2*q1, -2*q2, 2*q3
        
        # Gradient = J^T * f
        grad_q0 = j11*f1 + j21*f2 + j31*f3
        grad_q1 = j12*f1 + j22*f2 + j32*f3
        grad_q2 = j13*f1 + j23*f2 + j33*f3
        grad_q3 = j14*f1 + j24*f2 + j34*f3
        
        # Normalize gradient
        grad_norm = math.sqrt(grad_q0**2 + grad_q1**2 + grad_q2**2 + grad_q3**2)
        if grad_norm > 0.0001:
            grad_q0 /= grad_norm
            grad_q1 /= grad_norm
            grad_q2 /= grad_norm
            grad_q3 /= grad_norm
        
        # Quaternion derivative = gyroscope term - beta * gradient
        q0_dot = 0.5 * (-q1 * gx_rad - q2 * gy_rad - q3 * gz_rad) - self.beta * grad_q0
        q1_dot = 0.5 * (q0 * gx_rad + q2 * gz_rad - q3 * gy_rad) - self.beta * grad_q1
        q2_dot = 0.5 * (q0 * gy_rad - q1 * gz_rad + q3 * gx_rad) - self.beta * grad_q2
        q3_dot = 0.5 * (q0 * gz_rad + q1 * gy_rad - q2 * gx_rad) - self.beta * grad_q3
        
        # Integration update
        q0 = q0 + q0_dot * dt
        q1 = q1 + q1_dot * dt
        q2 = q2 + q2_dot * dt
        q3 = q3 + q3_dot * dt
        
        # Renormalize
        q_norm = math.sqrt(q0**2 + q1**2 + q2**2 + q3**2)
        self.q = [q0/q_norm, q1/q_norm, q2/q_norm, q3/q_norm]
        
        return self.q, self.initialized
    
    def reset(self):
        """Reset quaternion to unit quaternion"""
        self.q = [1.0, 0.0, 0.0, 0.0]
        self.initialized = False
        self.init_buffer.clear()

# IMU Data Buffer
class IMUDataBuffer:
    def __init__(self, max_size=1000):
        self.max_size = max_size
        self.accel_x = deque(maxlen=max_size)
        self.accel_y = deque(maxlen=max_size)
        self.accel_z = deque(maxlen=max_size)
        self.accel_motion_x = deque(maxlen=max_size)  # Motion Acceleration (Gravity-Removed)
        self.accel_motion_y = deque(maxlen=max_size)
        self.accel_motion_z = deque(maxlen=max_size)
        self.gyro_x = deque(maxlen=max_size)
        self.gyro_y = deque(maxlen=max_size)
        self.gyro_z = deque(maxlen=max_size)
        self.quat_w = deque(maxlen=max_size)  # Quaternion
        self.quat_x = deque(maxlen=max_size)
        self.quat_y = deque(maxlen=max_size)
        self.quat_z = deque(maxlen=max_size)
        self.timestamps = deque(maxlen=max_size)

    def add(self, accel_x, accel_y, accel_z, accel_motion_x, accel_motion_y, accel_motion_z, gyro_x, gyro_y, gyro_z, quat_w, quat_x, quat_y, quat_z, timestamp):
        self.accel_x.append(accel_x)
        self.accel_y.append(accel_y)
        self.accel_z.append(accel_z)
        self.accel_motion_x.append(accel_motion_x)
        self.accel_motion_y.append(accel_motion_y)
        self.accel_motion_z.append(accel_motion_z)
        self.gyro_x.append(gyro_x)
        self.gyro_y.append(gyro_y)
        self.gyro_z.append(gyro_z)
        self.quat_w.append(quat_w)
        self.quat_x.append(quat_x)
        self.quat_y.append(quat_y)
        self.quat_z.append(quat_z)
        self.timestamps.append(timestamp)

    def clear(self):
        self.accel_x.clear()
        self.accel_y.clear()
        self.accel_z.clear()
        self.accel_motion_x.clear()
        self.accel_motion_y.clear()
        self.accel_motion_z.clear()
        self.gyro_x.clear()
        self.gyro_y.clear()
        self.gyro_z.clear()
        self.quat_w.clear()
        self.quat_x.clear()
        self.quat_y.clear()
        self.quat_z.clear()
        self.timestamps.clear()

    def get_rows(self):
        return list(zip(self.timestamps, self.accel_x, self.accel_y, self.accel_z, self.gyro_x, self.gyro_y, self.gyro_z))

    def __len__(self):
        return len(self.timestamps)


# BLE Scanning and Connection Thread
class BLEThread(QThread):
    device_found = pyqtSignal(str, str)  # name, address
    connected = pyqtSignal(bool, str)  # success, message
    data_received = pyqtSignal(dict)  # IMU data
    disconnected = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.scan_mode = False
        self.connect_mode = False
        self.device_address = None
        self.client = None
        self.running = True

    def run(self):
        if self.scan_mode:
            asyncio.run(self.scan_devices())
        elif self.connect_mode:
            asyncio.run(self.connect_device())

    async def scan_devices(self):
        """Scan for Bluetooth devices"""
        self.device_found.emit("Scanning...", "")
        try:
            devices = await BleakScanner.discover()
            for device in devices:
                self.device_found.emit(device.name or "Unknown", device.address)
        except Exception as e:
            self.device_found.emit(f"Scan Failed: {str(e)}", "")
        self.device_found.emit("Scan complete", "")

    async def connect_device(self):
        """Connect to Bluetooth device and receive data"""
        disconnected_sent = False
        try:
            async with BleakClient(self.device_address) as client:
                self.client = client
                if client.is_connected:
                    self.connected.emit(True, f"Connected to {self.device_address}")

                    # Subscribe to IMU data characteristic
                    def notification_handler(sender, data):
                        if not self.running:
                            return
                        if len(data) >= 12:
                            # Parse 6 int16_t values
                            values = struct.unpack('<hhhhhh', data[:12])
                            imu_data = {
                                'accel_x': values[0],
                                'accel_y': values[1],
                                'accel_z': values[2],
                                'gyro_x': values[3],
                                'gyro_y': values[4],
                                'gyro_z': values[5],
                                'timestamp': datetime.now().strftime("%H:%M:%S.%f")[:-3]
                            }
                            self.data_received.emit(imu_data)

                    await client.start_notify(IMU_DATA_CHAR_UUID, notification_handler)

                    # Maintain connection
                    while self.connect_mode and self.running:
                        await asyncio.sleep(0.1)

                    # Normal disconnect
                    try:
                        await client.stop_notify(IMU_DATA_CHAR_UUID)
                    except Exception:
                        pass  # Ignore stop notify errors
                else:
                    self.connected.emit(False, "Connection Failed")
                    disconnected_sent = True
        except Exception as e:
            self.connected.emit(False, f"Error: {str(e)}")
        finally:
            # Ensure disconnect signal sent only once
            if not disconnected_sent:
                self.disconnected.emit()


    def stop(self):
        self.running = False
        self.scan_mode = False
        self.connect_mode = False


# 3D Quaternion Visualization Canvas
class Quaternion3DCanvas(FigureCanvas):
    """Display quaternion-based coordinate system rotation using matplotlib"""
    def __init__(self, parent=None):
        self.fig = Figure(figsize=(4, 5), dpi=100)
        self.ax = self.fig.add_subplot(111, projection='3d')
        super().__init__(self.fig)
        self.setParent(parent)
        self.fig.subplots_adjust(left=0.12, right=0.95, top=0.92, bottom=0.18)
        
        # 
        self.colors = ['red', 'green', 'blue']  # X, Y, Z
        self.axis_labels = ['X', 'Y', 'Z']
        
        # 
        self.ax.set_xlim([-1.5, 1.5])
        self.ax.set_ylim([-1.5, 1.5])
        self.ax.set_zlim([-1.5, 1.5])
        
        # 
        self.ax.set_xlabel('X')
        self.ax.set_ylabel('Y')
        self.ax.set_zlabel('Z')
        self.ax.set_title('Quaternion 3D', fontsize=10)
        
        # 
        self.ax.grid(True, alpha=0.3)
        
    def update_rotation(self, q):
        """
        Quaternion3D
        q: [w, x, y, z]
        """
        # Clear
        for collection in list(self.ax.collections):
            collection.remove()
        for line in list(self.ax.lines):
            line.remove()
        
        # Quaternion
        q0, q1, q2, q3 = q
        R = np.array([
            [1 - 2*(q2**2 + q3**2), 2*(q1*q2 - q0*q3), 2*(q1*q3 + q0*q2)],
            [2*(q1*q2 + q0*q3), 1 - 2*(q1**2 + q3**2), 2*(q2*q3 - q0*q1)],
            [2*(q1*q3 - q0*q2), 2*(q2*q3 + q0*q1), 1 - 2*(q1**2 + q2**2)]
        ])
        
        # 
        origin = np.array([0, 0, 0])
        axes = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])  # X, Y, Z
        
        # 
        rotated_axes = (R @ axes.T).T
        
        # 
        for i, (axis, color, label) in enumerate(zip(rotated_axes, self.colors, self.axis_labels)):
            self.ax.quiver(0, 0, 0, axis[0], axis[1], axis[2], 
                          color=color, arrow_length_ratio=0.1, linewidth=3, label=label)
        
        # （，）
        original_axes = axes
        for i, (axis, color) in enumerate(zip(original_axes, self.colors)):
            self.ax.quiver(0, 0, 0, axis[0]*0.7, axis[1]*0.7, axis[2]*0.7,
                          color=color, arrow_length_ratio=0.1, linewidth=1, 
                          linestyle='--', alpha=0.3)
        
        # （Z）
        self.ax.quiver(0, 0, 0, 0, 0, -1, color='purple', arrow_length_ratio=0.1, 
                      linewidth=2, linestyle=':', label='Gravity', alpha=0.7)
        
        self.draw()


# 
class IMUReceiverApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("BLE IMU Data Receiver - 24-Hour Collection")
        self.setGeometry(100, 100, 1600, 1000)
        self.ble_thread = None
        self.data_buffer = IMUDataBuffer()
        self.data_file = None
        self.is_recording = False
        self.is_connected = False
        self.chart_update_counter = 0
        self.packet_count = 0
        
        # Quaternion
        self.quat_filter = QuaternionFilter(beta=0.1)
        
        self.write_queue = queue.Queue(maxsize=240)
        self.write_thread = None
        self.write_stop_flag = False
        self.sampling_interval = BLE_OUTPUT_INTERVAL_S
        self.last_packet_monotonic = None
        
        self.init_ui()

    def init_ui(self):
        """UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QHBoxLayout()

        # Left side: Control Panel
        left_layout = QVBoxLayout()

        # Connection Control
        control_group = QGroupBox("Connection Control")
        control_layout = QGridLayout()

        self.device_combo = QComboBox()
        control_layout.addWidget(QLabel("Device List:"), 0, 0)
        control_layout.addWidget(self.device_combo, 0, 1)

        self.scan_btn = QPushButton("Scan Devices")
        self.scan_btn.clicked.connect(self.scan_devices)
        control_layout.addWidget(self.scan_btn, 1, 0)

        self.connect_btn = QPushButton("Connect Device")
        self.connect_btn.clicked.connect(self.connect_device)
        self.connect_btn.setEnabled(False)
        control_layout.addWidget(self.connect_btn, 1, 1)

        self.disconnect_btn = QPushButton("Disconnect")
        self.disconnect_btn.clicked.connect(self.disconnect_device)
        self.disconnect_btn.setEnabled(False)
        control_layout.addWidget(self.disconnect_btn, 2, 0, 1, 2)

        control_group.setLayout(control_layout)
        left_layout.addWidget(control_group)

        # Real-time Data Display
        display_group = QGroupBox("Real-Time Data")
        display_layout = QGridLayout()

        # Acceleration
        accel_font = QFont()
        accel_font.setPointSize(11)
        accel_font.setBold(True)

        display_layout.addWidget(QLabel("Acceleration (m/s²)"), 0, 0, 1, 2)
        self.accel_x_label = QLabel("X: 0")
        self.accel_x_label.setFont(accel_font)
        self.accel_y_label = QLabel("Y: 0")
        self.accel_y_label.setFont(accel_font)
        self.accel_z_label = QLabel("Z: 0")
        self.accel_z_label.setFont(accel_font)
        display_layout.addWidget(self.accel_x_label, 1, 0)
        display_layout.addWidget(self.accel_y_label, 1, 1)
        display_layout.addWidget(self.accel_z_label, 2, 0)

        # Motion Acceleration (Gravity-Removed)
        display_layout.addWidget(QLabel("Motion Acceleration (m/s²)"), 3, 0, 1, 2)
        self.accel_motion_x_label = QLabel("X: 0")
        self.accel_motion_x_label.setFont(accel_font)
        self.accel_motion_x_label.setStyleSheet("color: darkred;")
        self.accel_motion_y_label = QLabel("Y: 0")
        self.accel_motion_y_label.setFont(accel_font)
        self.accel_motion_y_label.setStyleSheet("color: darkgreen;")
        self.accel_motion_z_label = QLabel("Z: 0")
        self.accel_motion_z_label.setFont(accel_font)
        self.accel_motion_z_label.setStyleSheet("color: darkblue;")
        display_layout.addWidget(self.accel_motion_x_label, 4, 0)
        display_layout.addWidget(self.accel_motion_y_label, 4, 1)
        display_layout.addWidget(self.accel_motion_z_label, 5, 0)

        # Angular Velocity
        display_layout.addWidget(QLabel("Angular Velocity (°/s)"), 6, 0, 1, 2)
        self.gyro_x_label = QLabel("X: 0")
        self.gyro_x_label.setFont(accel_font)
        self.gyro_y_label = QLabel("Y: 0")
        self.gyro_y_label.setFont(accel_font)
        self.gyro_z_label = QLabel("Z: 0")
        self.gyro_z_label.setFont(accel_font)
        display_layout.addWidget(self.gyro_x_label, 7, 0)
        display_layout.addWidget(self.gyro_y_label, 7, 1)
        display_layout.addWidget(self.gyro_z_label, 8, 0)

        # Quaternion
        display_layout.addWidget(QLabel("Quaternion (W,X,Y,Z)"), 9, 0, 1, 2)
        self.quat_w_label = QLabel("W: 0")
        self.quat_w_label.setFont(accel_font)
        self.quat_w_label.setStyleSheet("color: purple;")
        self.quat_x_label = QLabel("X: 0")
        self.quat_x_label.setFont(accel_font)
        self.quat_x_label.setStyleSheet("color: red;")
        self.quat_y_label = QLabel("Y: 0")
        self.quat_y_label.setFont(accel_font)
        self.quat_y_label.setStyleSheet("color: green;")
        self.quat_z_label = QLabel("Z: 0")
        self.quat_z_label.setFont(accel_font)
        self.quat_z_label.setStyleSheet("color: blue;")
        display_layout.addWidget(self.quat_w_label, 10, 0)
        display_layout.addWidget(self.quat_x_label, 10, 1)
        display_layout.addWidget(self.quat_y_label, 11, 0)
        display_layout.addWidget(self.quat_z_label, 11, 1)

        # 
        display_layout.addWidget(QLabel("Timestamp:"), 12, 0)
        self.timestamp_label = QLabel("--:--:--.---")
        self.timestamp_label.setFont(accel_font)
        display_layout.addWidget(self.timestamp_label, 12, 1)

        # 
        display_layout.addWidget(QLabel("Packets Received:"), 13, 0)
        self.packet_count_label = QLabel("0")
        self.packet_count_label.setFont(accel_font)
        display_layout.addWidget(self.packet_count_label, 13, 1)

        display_group.setLayout(display_layout)
        left_layout.addWidget(display_group)

        # Data Recording
        record_group = QGroupBox("Data Recording")
        record_layout = QGridLayout()

        self.record_checkbox = QCheckBox("Enable Auto Recording (CSV)")
        self.record_checkbox.stateChanged.connect(self.toggle_recording)
        self.record_checkbox.setEnabled(False)
        record_layout.addWidget(self.record_checkbox, 0, 0, 1, 2)

        # Recording Status Display
        self.recording_status_label = QLabel("Recording Status: Not Started")
        self.recording_status_label.setStyleSheet("color: gray;")
        record_layout.addWidget(self.recording_status_label, 1, 0, 1, 2)

        # Quick start/stop buttons
        self.quick_record_btn = QPushButton("▶ Start Recording")
        self.quick_record_btn.clicked.connect(self.quick_start_recording)
        self.quick_record_btn.setEnabled(False)
        record_layout.addWidget(self.quick_record_btn, 2, 0)

        self.quick_stop_btn = QPushButton("⏹ Stop Recording")
        self.quick_stop_btn.clicked.connect(self.quick_stop_recording)
        self.quick_stop_btn.setEnabled(False)
        record_layout.addWidget(self.quick_stop_btn, 2, 1)

        self.save_btn = QPushButton("💾 Save Data (CSV)")
        self.save_btn.clicked.connect(self.save_data)
        self.save_btn.setEnabled(False)
        record_layout.addWidget(self.save_btn, 3, 0, 1, 2)

        self.export_excel_btn = QPushButton("📊 Export to Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_excel_btn.setEnabled(False)
        record_layout.addWidget(self.export_excel_btn, 4, 0, 1, 2)

        self.clear_btn = QPushButton("🗑 Clear Data")
        self.clear_btn.clicked.connect(self.clear_data)
        record_layout.addWidget(self.clear_btn, 5, 0, 1, 2)

        record_group.setLayout(record_layout)
        left_layout.addWidget(record_group)

        # Status Information
        info_group = QGroupBox("Status Information")
        info_layout = QVBoxLayout()
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: green; font-weight: bold;")
        info_layout.addWidget(self.status_label)

        # Logs and 3D visualization side by side
        log_and_3d_layout = QHBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(220)
        self.log_text.setMaximumWidth(400)
        log_and_3d_layout.addWidget(self.log_text)

        # ===== Quaternion 3D Visualization =====
        self.quat_3d_canvas = Quaternion3DCanvas(self)
        self.quat_3d_canvas.setMaximumHeight(220)
        self.quat_3d_canvas.setMaximumWidth(320)
        log_and_3d_layout.addWidget(self.quat_3d_canvas)

        info_layout.addLayout(log_and_3d_layout)

        info_group.setLayout(info_layout)
        left_layout.addWidget(info_group)

        left_layout.addStretch()

        # Right side: Scatter plots (Acceleration + Angular Velocity)
        right_layout = QVBoxLayout()

        # ===== Acceleration Chart =====
        self.accel_chart = QChart()
        self.accel_chart.setTitle("Acceleration Data (Scatter Plot)")
        self.accel_chart.setAnimationOptions(QChart.NoAnimation)
        self.accel_chart_view = QChartView(self.accel_chart)
        self.accel_chart_view.setRenderHint(QPainter.Antialiasing)

        # Acceleration scatter series (3 colors)
        self.accel_scatter_series = []
        accel_labels = ["Accel X (m/s²)", "Accel Y (m/s²)", "Accel Z (m/s²)"]
        accel_colors = [
            QColor(255, 0, 0),      # 
            QColor(0, 200, 0),      # 
            QColor(0, 0, 255),      # 
        ]

        for label, color in zip(accel_labels, accel_colors):
            scatter = QScatterSeries()
            scatter.setName(label)
            scatter.setMarkerSize(7)
            scatter.setPen(QPen(color, 1))
            scatter.setBrush(color)
            self.accel_scatter_series.append(scatter)
            self.accel_chart.addSeries(scatter)

        self.accel_chart.createDefaultAxes()
        accel_axes_x = self.accel_chart.axes(Qt.Horizontal)
        accel_axes_y = self.accel_chart.axes(Qt.Vertical)
        if accel_axes_x:
            accel_axes_x[0].setTitleText("Sample Index")
        if accel_axes_y:
            accel_axes_y[0].setTitleText("Acceleration (m/s²)")
        
        self.accel_chart.legend().setVisible(True)
        self.accel_chart.legend().setAlignment(Qt.AlignRight)

        right_layout.addWidget(self.accel_chart_view)

        # ===== Angular Velocity Chart =====
        self.gyro_chart = QChart()
        self.gyro_chart.setTitle("Angular Velocity Data (Scatter Plot)")
        self.gyro_chart.setAnimationOptions(QChart.NoAnimation)
        self.gyro_chart_view = QChartView(self.gyro_chart)
        self.gyro_chart_view.setRenderHint(QPainter.Antialiasing)

        # Angular velocity scatter series (3 colors)
        self.gyro_scatter_series = []
        gyro_labels = ["Gyro X (°/s)", "Gyro Y (°/s)", "Gyro Z (°/s)"]
        gyro_colors = [
            QColor(255, 165, 0),    # 
            QColor(128, 0, 128),    # 
            QColor(165, 42, 42),    # 
        ]

        for label, color in zip(gyro_labels, gyro_colors):
            scatter = QScatterSeries()
            scatter.setName(label)
            scatter.setMarkerSize(7)
            scatter.setPen(QPen(color, 1))
            scatter.setBrush(color)
            self.gyro_scatter_series.append(scatter)
            self.gyro_chart.addSeries(scatter)

        self.gyro_chart.createDefaultAxes()
        gyro_axes_x = self.gyro_chart.axes(Qt.Horizontal)
        gyro_axes_y = self.gyro_chart.axes(Qt.Vertical)
        if gyro_axes_x:
            gyro_axes_x[0].setTitleText("Sample Index")
        if gyro_axes_y:
            gyro_axes_y[0].setTitleText("Angular Velocity (°/s)")
        
        self.gyro_chart.legend().setVisible(True)
        self.gyro_chart.legend().setAlignment(Qt.AlignRight)

        right_layout.addWidget(self.gyro_chart_view)

        # ===== Motion Acceleration Chart =====
        self.motion_chart = QChart()
        self.motion_chart.setTitle("Motion Acceleration Data (Scatter Plot)")
        self.motion_chart.setAnimationOptions(QChart.NoAnimation)
        self.motion_chart_view = QChartView(self.motion_chart)
        self.motion_chart_view.setRenderHint(QPainter.Antialiasing)

        # Motion acceleration scatter series (3 colors)
        self.motion_scatter_series = []
        motion_labels = ["Motion X (m/s²)", "Motion Y (m/s²)", "Motion Z (m/s²)"]
        motion_colors = [
            QColor(200, 0, 0),      # 
            QColor(0, 150, 0),      # 
            QColor(0, 0, 200),      # 
        ]

        for label, color in zip(motion_labels, motion_colors):
            scatter = QScatterSeries()
            scatter.setName(label)
            scatter.setMarkerSize(7)
            scatter.setPen(QPen(color, 1))
            scatter.setBrush(color)
            self.motion_scatter_series.append(scatter)
            self.motion_chart.addSeries(scatter)

        self.motion_chart.createDefaultAxes()
        motion_axes_x = self.motion_chart.axes(Qt.Horizontal)
        motion_axes_y = self.motion_chart.axes(Qt.Vertical)
        if motion_axes_x:
            motion_axes_x[0].setTitleText("Sample Index")
        if motion_axes_y:
            motion_axes_y[0].setTitleText("Motion Acceleration (m/s²)")
        
        self.motion_chart.legend().setVisible(True)
        self.motion_chart.legend().setAlignment(Qt.AlignRight)

        right_layout.addWidget(self.motion_chart_view)

        # Overall Layout
        layout.addLayout(left_layout, 1)
        layout.addLayout(right_layout, 1)
        central_widget.setLayout(layout)

    def scan_devices(self):
        """Scan Devices"""
        # Stop old thread
        if self.ble_thread and self.ble_thread.isRunning():
            self.ble_thread.running = False
            self.ble_thread.stop()
            self.ble_thread.wait(2000)
        
        self.scan_btn.setEnabled(False)
        self.device_combo.clear()
        self.device_combo.addItem("Scanning...")
        self.log("Starting device scan...")

        self.ble_thread = BLEThread()
        self.ble_thread.device_found.connect(self.on_device_found)
        self.ble_thread.scan_mode = True
        self.ble_thread.start()

    def on_device_found(self, name, address):
        """Device discovery callback"""
        if "Scanning" in name:
            self.device_combo.clear()
            return

        if not address:
            self.scan_btn.setEnabled(True)
            if self.device_combo.count() > 0:
                self.connect_btn.setEnabled(True)
            self.log("Scan complete")
            return

        self.log(f"Device: {name} ({address})")
        self.device_combo.addItem(f"{name or 'Unknown'}", address)

    def connect_device(self):
        """Connect Device"""
        if self.device_combo.count() == 0:
            QMessageBox.warning(self, "Error", "Please scan devices first")
            return

        device_address = self.device_combo.currentData()
        device_name = self.device_combo.currentText()

        if not device_address:
            QMessageBox.warning(self, "Error", "Please select a valid device")
            return

        # Stop old thread
        if self.ble_thread and self.ble_thread.isRunning():
            self.ble_thread.running = False
            self.ble_thread.stop()
            self.ble_thread.wait(2000)

        self.connect_btn.setEnabled(False)
        self.scan_btn.setEnabled(False)
        self.log(f" {device_name}...")

        self.ble_thread = BLEThread()
        self.ble_thread.connected.connect(self.on_connected)
        self.ble_thread.data_received.connect(self.on_data_received)
        self.ble_thread.disconnected.connect(self.on_disconnected)
        self.ble_thread.device_address = device_address
        self.ble_thread.connect_mode = True
        self.ble_thread.start()

    def on_connected(self, success, message):
        """Connection callback"""
        self.log(message)
        if success:
            self.is_connected = True
            # Reset quaternion filter
            self.quat_filter.reset()
            self.last_packet_monotonic = None
            
            self.status_label.setText("Connected")
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
            self.disconnect_btn.setEnabled(True)
            self.record_checkbox.setEnabled(True)
            self.quick_record_btn.setEnabled(True)
            self.quick_stop_btn.setEnabled(True)
            self.save_btn.setEnabled(True)
            self.export_excel_btn.setEnabled(True)
        else:
            self.status_label.setText("Connection Failed")
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
            self.connect_btn.setEnabled(True)
            self.scan_btn.setEnabled(True)

    def on_data_received(self, imu_data):
        """Data reception callback"""
        now_monotonic = time.perf_counter()
        if self.last_packet_monotonic is None:
            sample_dt = self.sampling_interval
        else:
            sample_dt = now_monotonic - self.last_packet_monotonic
            sample_dt = min(max(sample_dt, 1.0 / 120.0), 0.1)
        self.last_packet_monotonic = now_monotonic
        # Convert firmware milli units to readable units: acceleration m/s², angular velocity °/s
        ax_ms2 = imu_data['accel_x'] / 1000.0
        ay_ms2 = imu_data['accel_y'] / 1000.0
        az_ms2 = imu_data['accel_z'] / 1000.0
        # Firmware sends mrad/s, convert to °/s: rad→deg (180/π) and divide by 1000
        to_dps = 0.0572957795
        gx_dps = imu_data['gyro_x'] * to_dps
        gy_dps = imu_data['gyro_y'] * to_dps
        gz_dps = imu_data['gyro_z'] * to_dps

        # Calculate quaternion using Madgwick filter
        q_values, is_initialized = self.quat_filter.update(
            ax_ms2, ay_ms2, az_ms2, gx_dps, gy_dps, gz_dps, dt=sample_dt
        )
        qw, qx, qy, qz = q_values
        
        # Calculate motion acceleration (gravity-removed)
        # Only calculate after initialization
        if is_initialized:
            accel_motion = self.quat_filter.remove_gravity([ax_ms2, ay_ms2, az_ms2])
            ax_motion, ay_motion, az_motion = accel_motion
        else:
            ax_motion, ay_motion, az_motion = 0, 0, 0

        # 
        self.accel_x_label.setText(f"X: {ax_ms2:.2f}")
        self.accel_y_label.setText(f"Y: {ay_ms2:.2f}")
        self.accel_z_label.setText(f"Z: {az_ms2:.2f}")
        
        # 
        if is_initialized:
            self.accel_motion_x_label.setText(f"X: {ax_motion:.2f}")
            self.accel_motion_y_label.setText(f"Y: {ay_motion:.2f}")
            self.accel_motion_z_label.setText(f"Z: {az_motion:.2f}")
        else:
            self.accel_motion_x_label.setText(f"X: ...")
            self.accel_motion_y_label.setText(f"Y: ...")
            self.accel_motion_z_label.setText(f"Z: ...")
        
        # 
        self.gyro_x_label.setText(f"X: {gx_dps:.2f}")
        self.gyro_y_label.setText(f"Y: {gy_dps:.2f}")
        self.gyro_z_label.setText(f"Z: {gz_dps:.2f}")
        
        # Quaternion
        self.quat_w_label.setText(f"W: {qw:.4f}")
        self.quat_x_label.setText(f"X: {qx:.4f}")
        self.quat_y_label.setText(f"Y: {qy:.4f}")
        self.quat_z_label.setText(f"Z: {qz:.4f}")
        
        self.timestamp_label.setText(imu_data['timestamp'])

        # 
        self.packet_count += 1
        self.packet_count_label.setText(str(self.packet_count))

        # Save data to buffer (including quaternion)
        self.data_buffer.add(
            ax_ms2, ay_ms2, az_ms2,
            ax_motion, ay_motion, az_motion,
            gx_dps, gy_dps, gz_dps,
            qw, qx, qy, qz,
            imu_data['timestamp']
        )

        # If recording enabled, put data in write queue (no direct write to avoid blocking)
        if self.is_recording:
            try:
                self.write_queue.put_nowait({
                    'timestamp': imu_data['timestamp'],
                    'accel_x': ax_ms2,
                    'accel_y': ay_ms2,
                    'accel_z': az_ms2,
                    'accel_motion_x': ax_motion,
                    'accel_motion_y': ay_motion,
                    'accel_motion_z': az_motion,
                    'gyro_x': gx_dps,
                    'gyro_y': gy_dps,
                    'gyro_z': gz_dps,
                    'quat_w': qw,
                    'quat_x': qx,
                    'quat_y': qy,
                    'quat_z': qz,
                })
            except queue.Full:
                # Queue full means write thread falling behind, log warning but don't discard
                self.log("⚠️  Write queue full, possible data loss!")

        # Update charts (every 5 data points to avoid lag)
        self.chart_update_counter += 1
        if self.chart_update_counter >= 5:
            self.chart_update_counter = 0
            self.update_charts()
            # Update 3D quaternion visualization (every 10 data points)
            if self.packet_count % 10 == 0:
                self.quat_3d_canvas.update_rotation(self.quat_filter.q)

    def update_charts(self):
        """Update charts"""
        # Get data length
        data_len = len(self.data_buffer)
        if data_len == 0:
            return

        # Only display last 200 data points (avoid lag)
        max_display_points = 200
        start_idx = max(0, data_len - max_display_points)

        # Clear all acceleration scatter points
        for series in self.accel_scatter_series:
            series.clear()

        # Clear all angular velocity scatter points
        for series in self.gyro_scatter_series:
            series.clear()

        # Clear all motion acceleration scatter points
        for series in self.motion_scatter_series:
            series.clear()

        # Get data (only recent part)
        accel_x_data = list(self.data_buffer.accel_x)[start_idx:]
        accel_y_data = list(self.data_buffer.accel_y)[start_idx:]
        accel_z_data = list(self.data_buffer.accel_z)[start_idx:]
        motion_x_data = list(self.data_buffer.accel_motion_x)[start_idx:]
        motion_y_data = list(self.data_buffer.accel_motion_y)[start_idx:]
        motion_z_data = list(self.data_buffer.accel_motion_z)[start_idx:]
        gyro_x_data = list(self.data_buffer.gyro_x)[start_idx:]
        gyro_y_data = list(self.data_buffer.gyro_y)[start_idx:]
        gyro_z_data = list(self.data_buffer.gyro_z)[start_idx:]

        # Fill acceleration scatter (using global index as X-axis)
        accel_data = [accel_x_data, accel_y_data, accel_z_data]
        for series_idx, data_values in enumerate(accel_data):
            for local_idx, y_val in enumerate(data_values):
                global_x = start_idx + local_idx
                self.accel_scatter_series[series_idx].append(QPointF(global_x, y_val))

        # Fill motion acceleration scatter (using global index as X-axis)
        motion_data = [motion_x_data, motion_y_data, motion_z_data]
        for series_idx, data_values in enumerate(motion_data):
            for local_idx, y_val in enumerate(data_values):
                global_x = start_idx + local_idx
                self.motion_scatter_series[series_idx].append(QPointF(global_x, y_val))

        # Fill angular velocity scatter (using global index as X-axis)
        gyro_data = [gyro_x_data, gyro_y_data, gyro_z_data]
        for series_idx, data_values in enumerate(gyro_data):
            for local_idx, y_val in enumerate(data_values):
                global_x = start_idx + local_idx
                self.gyro_scatter_series[series_idx].append(QPointF(global_x, y_val))

        # Dynamically adjust axis ranges
        if len(accel_x_data) > 0:
            # Calculate acceleration range
            all_accel_values = accel_x_data + accel_y_data + accel_z_data
            accel_min = min(all_accel_values)
            accel_max = max(all_accel_values)
            accel_range = accel_max - accel_min
            accel_margin = accel_range * 0.1 if accel_range > 0 else 1.0

            # Update acceleration chart axes
            accel_axes_x = self.accel_chart.axes(Qt.Horizontal)
            accel_axes_y = self.accel_chart.axes(Qt.Vertical)
            if accel_axes_x:
                accel_axes_x[0].setRange(start_idx, start_idx + len(accel_x_data))
            if accel_axes_y:
                accel_axes_y[0].setRange(accel_min - accel_margin, accel_max + accel_margin)

            # Calculate motion acceleration range
            all_motion_values = motion_x_data + motion_y_data + motion_z_data
            if len(all_motion_values) > 0:
                motion_min = min(all_motion_values)
                motion_max = max(all_motion_values)
                motion_range = motion_max - motion_min
                motion_margin = motion_range * 0.1 if motion_range > 0 else 0.1

                # Update motion acceleration chart axes
                motion_axes_x = self.motion_chart.axes(Qt.Horizontal)
                motion_axes_y = self.motion_chart.axes(Qt.Vertical)
                if motion_axes_x:
                    motion_axes_x[0].setRange(start_idx, start_idx + len(motion_x_data))
                if motion_axes_y:
                    motion_axes_y[0].setRange(motion_min - motion_margin, motion_max + motion_margin)

            # Calculate angular velocity range
            all_gyro_values = gyro_x_data + gyro_y_data + gyro_z_data
            gyro_min = min(all_gyro_values)
            gyro_max = max(all_gyro_values)
            gyro_range = gyro_max - gyro_min
            gyro_margin = gyro_range * 0.1 if gyro_range > 0 else 1.0

            # Update angular velocity chart axes
            gyro_axes_x = self.gyro_chart.axes(Qt.Horizontal)
            gyro_axes_y = self.gyro_chart.axes(Qt.Vertical)
            if gyro_axes_x:
                gyro_axes_x[0].setRange(start_idx, start_idx + len(gyro_x_data))
            if gyro_axes_y:
                gyro_axes_y[0].setRange(gyro_min - gyro_margin, gyro_max + gyro_margin)

    def toggle_recording(self, state):
        """Toggle recording state"""
        if state:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"imu_data_{timestamp}.csv"
            try:
                self.data_file = open(filename, 'w', buffering=1)  # Line buffering
                self.data_file.write("timestamp,accel_x,accel_y,accel_z,accel_motion_x,accel_motion_y,accel_motion_z,gyro_x,gyro_y,gyro_z,quat_w,quat_x,quat_y,quat_z\n")
                self.is_recording = True
                self.write_stop_flag = False
                
                # Start dedicated write thread at the BLE output rate.
                self.write_thread = FileWriteThread(
                    self.write_queue,
                    self.data_file,
                    self.sampling_interval,
                    self
                )
                self.write_thread.write_error.connect(self.on_write_error)
                self.write_thread.start()
                
                self.recording_status_label.setText(f"✓ Record: {filename}")
                self.recording_status_label.setStyleSheet("color: green; font-weight: bold;")
                self.log(f"✓ StartRecord: {filename}")
                self.log(f"Rate: {BLE_OUTPUT_HZ:.0f} Hz BLE | {IMU_SENSOR_ODR_HZ:.0f} Hz IMU ODR")
            except Exception as e:
                QMessageBox.critical(self, "Error", f": {str(e)}")
                self.record_checkbox.setChecked(False)
                self.is_recording = False
        else:
            self.is_recording = False
            self.write_stop_flag = True
            
            # Wait for write thread to finish
            if self.write_thread and self.write_thread.isRunning():
                self.write_thread.wait()
            
            if self.data_file:
                self.data_file.flush()
                self.data_file.close()
                self.data_file = None
            
            self.recording_status_label.setText("Recording Status: Stopped")
            self.recording_status_label.setStyleSheet("color: gray;")
            self.log("✓ Auto recording stopped")

    def quick_start_recording(self):
        """Quick start recording"""
        if not self.is_recording:
            self.record_checkbox.setChecked(True)

    def quick_stop_recording(self):
        """Quick stop recording"""
        if self.is_recording:
            self.record_checkbox.setChecked(False)

    def save_data(self):
        """Save data to CSV"""
        filename, _ = QFileDialog.getSaveFileName(self, "Save Data", "", "CSV Files (*.csv)")
        if filename:
            try:
                with open(filename, 'w') as f:
                    f.write("timestamp,accel_x,accel_y,accel_z,accel_motion_x,accel_motion_y,accel_motion_z,gyro_x,gyro_y,gyro_z,quat_w,quat_x,quat_y,quat_z\n")
                    for ts, ax, ay, az, ax_m, ay_m, az_m, gx, gy, gz, qw, qx, qy, qz in zip(
                        self.data_buffer.timestamps,
                        self.data_buffer.accel_x, self.data_buffer.accel_y, self.data_buffer.accel_z,
                        self.data_buffer.accel_motion_x, self.data_buffer.accel_motion_y, self.data_buffer.accel_motion_z,
                        self.data_buffer.gyro_x, self.data_buffer.gyro_y, self.data_buffer.gyro_z,
                        self.data_buffer.quat_w, self.data_buffer.quat_x, self.data_buffer.quat_y, self.data_buffer.quat_z
                    ):
                        f.write(f"{ts},{ax:.5f},{ay:.5f},{az:.5f},{ax_m:.5f},{ay_m:.5f},{az_m:.5f},{gx:.5f},{gy:.5f},{gz:.5f},{qw:.6f},{qx:.6f},{qy:.6f},{qz:.6f}\n")
                self.log(f"✓ CSV Save: {filename}")
                QMessageBox.information(self, "Success", f"Save:\n{filename}\n\n {len(self.data_buffer)} Record")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"SaveFailed: {str(e)}")

    def export_to_excel(self):
        """Export data to Excel"""
        filename, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "", "Excel Files (*.xlsx)")
        if filename:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "IMU"

                # Set headers (including quaternion)
                headers = ["Index", "Accel X (m/s²)", "Accel Y (m/s²)", "Accel Z (m/s²)",
                          "Motion Accel X (m/s²)", "Motion Accel Y (m/s²)", "Motion Accel Z (m/s²)",
                          "Gyro X (°/s)", "Gyro Y (°/s)", "Gyro Z (°/s)",
                          "Quat W", "Quat X", "Quat Y", "Quat Z"]
                ws.append(headers)

                # Header formatting
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Fill data (including quaternion)
                for i, (ax, ay, az, ax_m, ay_m, az_m, gx, gy, gz, qw, qx, qy, qz) in enumerate(zip(
                    self.data_buffer.accel_x, self.data_buffer.accel_y, self.data_buffer.accel_z,
                    self.data_buffer.accel_motion_x, self.data_buffer.accel_motion_y, self.data_buffer.accel_motion_z,
                    self.data_buffer.gyro_x, self.data_buffer.gyro_y, self.data_buffer.gyro_z,
                    self.data_buffer.quat_w, self.data_buffer.quat_x, self.data_buffer.quat_y, self.data_buffer.quat_z
                )):
                    ws.append([i, f"{ax:.5f}", f"{ay:.5f}", f"{az:.5f}",
                              f"{ax_m:.5f}", f"{ay_m:.5f}", f"{az_m:.5f}",
                              f"{gx:.5f}", f"{gy:.5f}", f"{gz:.5f}",
                              f"{qw:.6f}", f"{qx:.6f}", f"{qy:.6f}", f"{qz:.6f}"])

                # Auto-adjust column width
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(filename)
                self.log(f"✓ Excel Export: {filename}")
                QMessageBox.information(self, "Success", f"Export Excel:\n{filename}\n\n {len(self.data_buffer)} Record")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"ExportFailed: {str(e)}")
                self.log(f"❌ Excel ExportFailed: {str(e)}")

    def clear_data(self):
        """Clear data"""
        reply = QMessageBox.question(self, "Confirm", "Are you sure you want to clear all buffered data?\n(Data already recorded in files will not be deleted)", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.data_buffer.clear()
            self.packet_count = 0
            self.packet_count_label.setText("0")
            self.update_charts()
            self.log("✓ Buffered data cleared")

    def disconnect_device(self):
        """Disconnect"""
        if self.ble_thread:
            self.log("Disconnecting...")
            self.ble_thread.connect_mode = False
            self.ble_thread.running = False
            self.ble_thread.stop()
            # Wait for thread to fully stop
            if self.ble_thread.isRunning():
                self.ble_thread.wait(2000)  # Don't use timeout= keyword
            # Don't set to None here, let on_disconnected handle it
            # self.ble_thread = None

    def on_disconnected(self):
        """Disconnection callback"""
        # Prevent duplicate processing
        if not self.is_connected:
            return
        
        self.is_connected = False
        self.last_packet_monotonic = None
        self.status_label.setText("Disconnected")
        self.status_label.setStyleSheet("color: red; font-weight: bold;")
        self.disconnect_btn.setEnabled(False)
        self.record_checkbox.setEnabled(False)
        self.record_checkbox.setChecked(False)
        self.quick_record_btn.setEnabled(False)
        self.quick_stop_btn.setEnabled(False)
        self.connect_btn.setEnabled(True)
        self.scan_btn.setEnabled(True)
        self.recording_status_label.setText("Recording Status: Not Started")
        self.recording_status_label.setStyleSheet("color: gray;")
        
        # Clean up thread reference
        if self.ble_thread:
            self.ble_thread = None
        
        self.log("Device disconnected")

    def on_write_error(self, error_msg):
        """File write error callback"""
        self.log(f"❌ Error: {error_msg}")
        QMessageBox.warning(self, "Write Error", error_msg)

    def log(self, message):
        """Add log message"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")

    def closeEvent(self, event):
        """Close event"""
        if self.ble_thread:
            self.ble_thread.running = False
            self.ble_thread.stop()
            if self.ble_thread.isRunning():
                self.ble_thread.wait(2000)
        
        # Stop recording
        if self.is_recording:
            self.record_checkbox.setChecked(False)
        
        # Ensure write thread has stopped
        if self.write_thread and self.write_thread.isRunning():
            self.write_stop_flag = True
            self.write_thread.wait(2000)
        
        if self.data_file:
            self.data_file.close()
        
        event.accept()


# Dedicated file write thread paced to the BLE output rate.
class FileWriteThread(QThread):
    write_error = pyqtSignal(str)
    
    def __init__(self, queue, file_obj, interval, parent):
        super().__init__()
        self.queue = queue
        self.file_obj = file_obj
        self.interval = interval
        self.parent = parent
        
    def run(self):
        """Thread main loop: periodically read one data point from the queue and write it."""
        last_write_time = time.time()
        write_count = 0
        
        try:
            while not self.parent.write_stop_flag:
                current_time = time.time()
                time_since_last = current_time - last_write_time
                
                if time_since_last >= self.interval:
                    # Time's up, try to get one data point from queue
                    try:
                        data = self.queue.get(timeout=0.01)
                        self.file_obj.write(
                            f"{data['timestamp']},{data['accel_x']:.5f},"
                            f"{data['accel_y']:.5f},{data['accel_z']:.5f},"
                            f"{data['accel_motion_x']:.5f},{data['accel_motion_y']:.5f},"
                            f"{data['accel_motion_z']:.5f},"
                            f"{data['gyro_x']:.5f},{data['gyro_y']:.5f},"
                            f"{data['gyro_z']:.5f},"
                            f"{data['quat_w']:.6f},{data['quat_x']:.6f},"
                            f"{data['quat_y']:.6f},{data['quat_z']:.6f}\n"
                        )
                        write_count += 1
                        last_write_time = current_time
                    except queue.Empty:
                        # Queue empty but time's up, data may still be in transit, keep waiting
                        pass
                else:
                    # Not yet, short sleep
                    time.sleep(0.001)
            
            # Before exiting, flush remaining data in queue
            while True:
                try:
                    data = self.queue.get_nowait()
                    self.file_obj.write(
                        f"{data['timestamp']},{data['accel_x']:.5f},"
                        f"{data['accel_y']:.5f},{data['accel_z']:.5f},"
                        f"{data['accel_motion_x']:.5f},{data['accel_motion_y']:.5f},"
                        f"{data['accel_motion_z']:.5f},"
                        f"{data['gyro_x']:.5f},{data['gyro_y']:.5f},"
                        f"{data['gyro_z']:.5f},"
                        f"{data['quat_w']:.6f},{data['quat_x']:.6f},"
                        f"{data['quat_y']:.6f},{data['quat_z']:.6f}\n"
                    )
                    write_count += 1
                except queue.Empty:
                    break
            
            # Final buffer flush
            self.file_obj.flush()
            
        except Exception as e:
            self.write_error.emit(f"Failed: {str(e)}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = IMUReceiverApp()
    window.show()
    sys.exit(app.exec_())
